import PySimpleGUI as sg
import pandas as pd
import os
from datetime import date
import datetime
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Border, Side, Alignment
from openpyxl.drawing.image import Image
import io
import urllib3
import os.path
from tabulate import tabulate
import statistics
import matplotlib.patches as mpatches
from matplotlib import pyplot as plt
import base64
import shutil

sg.theme('Light Green')

def DIW(y1,z1):
    data = pd.ExcelFile(workingpath)
    dfa = pd.DataFrame(pd.read_excel(data, 'DIW'))
    
    dfa['Date']= pd.to_datetime(dfa['Date']) 
    dfa['Date']= dfa['Date'].dt.date
    dfaa = dfa[(dfa['Date'] >= y1)]
    df = dfaa[(dfaa['Date'] <= z1)]
    
    #df.to_excel(r'C:\Users\HPP Assay\Desktop\Dump\df.xlsx', sheet_name='1')
    #return
    
    cbar = df.plot.bar(x='Date', y='counts', rot = 75, color = 'blue', figsize=(15,5))
    plt.tick_params(axis='x', labelsize=10)
    plt.setp(cbar.get_xticklabels()[1::2], visible=False)
    plt.tick_params(bottom=True, top=False, left=True, right=True)
    plt.ylabel('Cumulative Counts/mL')
    plt.axhline(y=50, color='r', linestyle='-')
    plt.xlabel('Date Range: ' + str(Date1a) + ' - ' + str(Date2a))
    plt.title('Daily DIW Quality at Startup')
    yellow_patch = mpatches.Patch(color='red', label='Ideal Counts <50')
    purple_patch = mpatches.Patch(color='blue', label='Cumulative Counts')
    plt.legend(handles=[yellow_patch, purple_patch])
    
    pathD = os.path.join(pathDIW, 'DIW plot ' + str(Date1a) + '-' + str(Date2a) + '.png')
    plt.savefig(pathD, bbox_inches='tight')

def convert_to_bytes(file_or_bytes, resize=None):
    import PIL.Image
    if isinstance(file_or_bytes, str):
        img = PIL.Image.open(file_or_bytes)
    else:
        try:
            img = PIL.Image.open(io.BytesIO(base64.b64decode(file_or_bytes)))
        except Exception:
            dataBytesIO = io.BytesIO(file_or_bytes)
            img = PIL.Image.open(dataBytesIO)

    cur_width, cur_height = img.size
    if resize:
        new_width, new_height = resize
        scale = min(new_height/cur_height, new_width/cur_width)
        img = img.resize((int(cur_width*scale), int(cur_height*scale)), PIL.Image.ANTIALIAS)
    bio = io.BytesIO()
    img.save(bio, format="PNG")
    del img
    return bio.getvalue()

def report(y):
#generate reported data log
    dafa= pd.DataFrame(pd.read_excel(pd.ExcelFile(workingpath), 'InLabLogdata'))
    dafab = pd.DataFrame(pd.read_excel(pd.ExcelFile(workingpath), 'InLabLogdata'))
    nan_value = float('NaN')
    dafa.replace('', nan_value, inplace=True)
    dafa.dropna(subset = ['Report'], inplace=True)
    dafab.dropna(subset = ['Product'], inplace=True)
    dafa['Report'].replace('Z','X', inplace = True)
    dafab['Report'].replace('Z','X', inplace = True)
    dafab['Report'].replace('Z*','Y', inplace = True)
    is_reported = dafa['Report']=='X'
    plotdf = dafa[is_reported]
    plotdfb = dafab[dafab.Report != 'Y']
    Path4b = os.path.join(Path2, 'data_all.xlsx')
    Path4 = os.path.join(Path2, 'data_reported.xlsx')
    plotdfb.to_excel(Path4b)
    plotdf.to_excel(Path4)
    
#Buld DF with only data to report        
    data = pd.ExcelFile(workingpath)
    dfa = pd.DataFrame(pd.read_excel(data, 'InLabLogdata'))
    
    dfaa = dfa[['Report', 'RP', 'Lot', 'PN', 'Product', 'Panalyst', 'PDT', 'CSid',
               'P02um', 'P03um', 'P05um','P10um', 'P20um', 'CS']]
    dfab = dfa[['Report', 'RA', 'Lot', 'Aanalyst', 'ADT', 'AC1', 'AC2', 'AC3',
               'A1', 'A1SD', 'A2', 'A2SD','A3', 'A3SD']]
    dfac = dfa[['Report', 'RO', 'Lot', 'Oanalyst', 'ODT', 'Oanalysis',
               'OpH', 'O1pH', 'OSG', 'OST']]
    dfad = dfa[['Report', 'RM', 'Lot', 'Manalyst', 'MDT', 'Mmethod',
               'MAl', 'MSb', 'MAs','MBa','MBe', 'MBi', 'MB', 'MCd', 
                'MCa','MCr','MCo','MCu','MGa','MGe','MAu','MFe','MPb','MLi',
                'MMg','MMn','MMo','MNb','MNi','MK','MAg','MNa','MSr','MTa',
                'MTl','MSn','MTi','MV','MZn','MZr']]
    
    dfza = dfaa[dfaa['Report']=='Z']
    dfz=dfza[dfza['RP']=='X']
    dfz.drop(columns=['Report'], inplace=True)
    dfba = dfab[dfab['Report']=='Z']
    dfb=dfba[dfba['RA']=='X']
    dfb.drop(columns=['Report'], inplace =True)
    dfca = dfac[dfac['Report']=='Z']
    dfc=dfca[dfca['RO']== 'X']
    dfc.drop(columns=['Report'], inplace=True)
    dfda = dfad[dfad['Report']=='Z']
    dfd = dfda[dfda['RM']=='X']
    dfd.drop(columns=['Report'], inplace=True)
    
    m1 = pd.merge(dfz, dfb, on='Lot', how="left")
    m2 = pd.merge(m1, dfc, on='Lot', how="left")
    df = pd.merge(m2, dfd, on='Lot', how="left")
    
    try:
        df['PDT']= pd.to_datetime(df['PDT'])
    except:
        sg.popup('Check Particle Test Date Formating in "Lab Data" sheet.')
        return
    try:
        df['ADT']= pd.to_datetime(df['ADT'])
    except:
        sg.popup('Check Assay Test Date Formating in "Lab Data" sheet.')
        return
    try:
        df['ODT']= pd.to_datetime(df['ODT'])
    except:
        sg.popup('Check ST/SG/Other Test Date Formating in "Lab Data" sheet.')
        return
    try:
        df['MDT']= pd.to_datetime(df['MDT'])
    except:
        sg.popup('Check Metals Test Date Formating in "Lab Data" sheet.')
        return
        
    #df.to_excel(r'C:\Users\HPP Assay\Desktop\Dump\df.xlsx', sheet_name='1')
    
#Import HPP logo image
    r = 1
    http = urllib3.PoolManager()
    r = http.request('GET', 'http://www.highpurityproducts.com/wp-content/uploads/2015/04/hpp-logo-main.png')
    image_file = io.BytesIO(r.data)

    img = Image(image_file)
    img.width = 230
    img.height = 60

#Export DF data to individual arrays for each unique report   
    zuples = [tuple(x) for x in df.to_numpy()]
    
#Format and create report    
    for zuple in zuples:
#create file
        wb = openpyxl.Workbook() 
        RF1076 = wb.active 
        RF1076.title = "RF-1076 HPP Data Report"
#cell size formatting    
        RF1076.column_dimensions['A'].width = 5.75
        RF1076.column_dimensions['B'].width = 14.6
        RF1076.column_dimensions['C'].width = 17.25
        RF1076.column_dimensions['D'].width = 10
        RF1076.column_dimensions['E'].width = 12
        RF1076.column_dimensions['F'].width = 12
        RF1076.column_dimensions['G'].width = 16.15
        RF1076.column_dimensions['H'].width = 5.75
            
        RF1076.row_dimensions[1].height = 50
        RF1076.row_dimensions[2].height = 30
        RF1076.row_dimensions[3].height = 20
        RF1076.row_dimensions[4].height = 16.5
        RF1076.row_dimensions[5].height = 9.5
        
#format styles        
        thin = Side(border_style="thin", color="000000")
        medium = Side(border_style="medium", color="000000")
        double = Side(border_style="double", color="000000")
        font1= Font(name='Times New Roman', size=12, 
                    bold=False, italic=False,
                    underline='none',strike=False,color='000000')
#Cell merges        
        RF1076.merge_cells('A1:F1')
        RF1076.merge_cells('G1:H1')
        RF1076.merge_cells('A2:F2')
        RF1076.merge_cells('G2:H2')
        RF1076.merge_cells('A3:F3')
        RF1076.merge_cells('G3:H3')
        RF1076.merge_cells('A4:H4')
        RF1076.merge_cells('D9:G9')
        RF1076.merge_cells('F10:G10')
        RF1076.merge_cells('F11:G13')
        RF1076.merge_cells('D10:E10')
        RF1076.merge_cells('D11:E11')
        RF1076.merge_cells('D12:E12')
        RF1076.merge_cells('D13:E13')
        
#Border Control        
        RF1076['A1'].border = Border(top=medium, left=medium, bottom=medium)
        RF1076['B1'].border = Border(top=medium, bottom=medium)
        RF1076['C1'].border = Border(top=medium, bottom=medium)
        RF1076['D1'].border = Border(top=medium, bottom=medium)
        RF1076['E1'].border = Border(top=medium, bottom=medium)
        RF1076['F1'].border = Border(top=medium, bottom=medium, right=medium)
        RF1076['G1'].border = Border(top=medium, bottom=medium)
        RF1076['H1'].border = Border(top=medium, right=medium, bottom=medium)
        
        RF1076['A2'].border = Border(left=medium, bottom=double)
        RF1076['B2'].border = Border(bottom=double)
        RF1076['C2'].border = Border(bottom=double)
        RF1076['D2'].border = Border(bottom=double)
        RF1076['E2'].border = Border(bottom=double)
        RF1076['F2'].border = Border(bottom=double, right=medium)
        RF1076['G2'].border = Border(bottom=double)
        RF1076['H2'].border = Border(right=medium, bottom=double)
        
        RF1076['A3'].border = Border(left=double, bottom=double)
        RF1076['B3'].border = Border(bottom=double)
        RF1076['C3'].border = Border(bottom=double)
        RF1076['D3'].border = Border(bottom=double)
        RF1076['E3'].border = Border(bottom=double)
        RF1076['F3'].border = Border(bottom=double, right=medium)
        RF1076['G3'].border = Border(bottom=double)
        RF1076['H3'].border = Border(right=double, bottom=double)
        
        RF1076['A4'].border = Border(left=double, bottom=double)
        RF1076['B4'].border = Border(bottom=double)
        RF1076['C4'].border = Border(bottom=double)
        RF1076['D4'].border = Border(bottom=double)
        RF1076['E4'].border = Border(bottom=double)
        RF1076['F4'].border = Border(bottom=double)
        RF1076['G4'].border = Border(bottom=double)
        RF1076['H4'].border = Border(right=double, bottom=double)
         
        RF1076['C6'].border = Border(bottom=thin)
        RF1076['C7'].border = Border(bottom=thin)
        RF1076['F6'].border = Border(bottom=thin)
        RF1076['F7'].border = Border(bottom=thin)
        
        RF1076['B9'].border = Border(top=medium, left= medium, bottom=thin)
        RF1076['B10'].border = Border(left= medium, bottom=thin)
        RF1076['B11'].border = Border(left= medium, bottom=thin)
        RF1076['B12'].border = Border(left= medium, bottom=thin)
        RF1076['B13'].border = Border(left= medium, bottom=medium)
        RF1076['B14'].border = Border(left= medium, bottom=medium)
        RF1076['B15'].border = Border(left= medium, right=thin)
        RF1076['B16'].border = Border(left= medium, right=thin)
        RF1076['B17'].border = Border(left= medium, right=thin)
        RF1076['B18'].border = Border(left= medium, right=thin)
        RF1076['B19'].border = Border(left= medium, right=thin, bottom=medium)
        RF1076['B20'].border = Border(left= medium, right=thin)
        RF1076['B21'].border = Border(left= medium, right=thin)
        RF1076['B22'].border = Border(left= medium, right=thin)
        RF1076['B23'].border = Border(left= medium, right=thin)
        RF1076['B24'].border = Border(left= medium, right=thin, bottom=medium)
        RF1076['B25'].border = Border(left= medium, right=thin)
        RF1076['B26'].border = Border(left= medium, right=thin)
        RF1076['B27'].border = Border(left= medium, right=thin)
        RF1076['B28'].border = Border(left= medium, right=thin)
        RF1076['B29'].border = Border(left= medium, right=thin, bottom=medium)
        RF1076['B30'].border = Border(left= medium, right=thin)
        RF1076['B31'].border = Border(left= medium, right=thin)
        RF1076['B32'].border = Border(left= medium, right=thin)
        RF1076['B33'].border = Border(left= medium, right=thin)
        RF1076['B34'].border = Border(left= medium, right=thin)
        RF1076['B35'].border = Border(left= medium, right=thin)
        RF1076['B36'].border = Border(left= medium, right=thin)
        RF1076['B37'].border = Border(left= medium, right=thin)
        RF1076['B38'].border = Border(left= medium, right=thin)
        RF1076['B39'].border = Border(left= medium, right=thin)
        RF1076['B40'].border = Border(left= medium, right=thin)
        RF1076['B41'].border = Border(left= medium, right=thin)
        RF1076['B42'].border = Border(left= medium, right=thin)
        RF1076['B43'].border = Border(left= medium, right=thin)
        RF1076['B44'].border = Border(left= medium, right=thin)
        RF1076['B45'].border = Border(left= medium, right=thin)
        RF1076['B46'].border = Border(left= medium, right=thin)
        RF1076['B47'].border = Border(left= medium, right=thin)
        RF1076['B48'].border = Border(left= medium, right=thin)
        RF1076['B49'].border = Border(left= medium, right=thin)
        RF1076['B50'].border = Border(left= medium, right=thin)
        RF1076['B51'].border = Border(left= medium, right=thin)
        RF1076['B52'].border = Border(left= medium, right=thin)
        RF1076['B53'].border = Border(left= medium, right=thin)
        RF1076['B54'].border = Border(left= medium, right=thin)
        RF1076['B55'].border = Border(left= medium, right=thin)
        RF1076['B56'].border = Border(left= medium, right=thin)
        RF1076['B57'].border = Border(left= medium, right=thin)
        RF1076['B58'].border = Border(left= medium, right=thin)
        RF1076['B59'].border = Border(left= medium, right=thin)
        RF1076['B60'].border = Border(left= medium, right=thin)
        RF1076['B61'].border = Border(left= medium, right=thin)
        RF1076['B62'].border = Border(left= medium, right=thin)
        RF1076['B63'].border = Border(left= medium, right=thin, bottom= medium)
        
        RF1076['C9'].border = Border(top=medium, bottom=thin, right=thin)
        RF1076['C10'].border = Border(bottom=thin, right=thin)
        RF1076['C11'].border = Border(bottom=thin, right=thin)
        RF1076['C12'].border = Border(bottom=thin, right=thin)
        RF1076['C13'].border = Border(bottom=medium, right=thin)
        RF1076['C14'].border = Border(bottom=medium, right=thin)
        RF1076['C19'].border = Border(bottom=medium)
        RF1076['C24'].border = Border(bottom=medium)
        RF1076['C29'].border = Border(bottom=medium)
        RF1076['C63'].border = Border(bottom= medium)

        RF1076['D9'].border = Border(top=medium, bottom=thin)
        RF1076['D10'].border = Border(bottom=thin)
        RF1076['D11'].border = Border(bottom=thin)
        RF1076['D12'].border = Border(bottom=thin)
        RF1076['D13'].border = Border(bottom=medium)
        RF1076['D14'].border = Border(bottom=medium, right=thin)
        RF1076['D19'].border = Border(bottom=medium)
        RF1076['D24'].border = Border(bottom=medium)
        RF1076['D29'].border = Border(bottom=medium)
        RF1076['D63'].border = Border(bottom= medium)

        RF1076['E9'].border = Border(top=medium, bottom=thin)
        RF1076['E10'].border = Border(bottom=thin, right=thin)
        RF1076['E11'].border = Border(bottom=thin, right=thin)
        RF1076['E12'].border = Border(bottom=thin, right=thin)
        RF1076['E13'].border = Border(bottom=medium, right=thin)
        RF1076['E14'].border = Border(bottom=medium, right=thin)
        RF1076['E19'].border = Border(bottom=medium)
        RF1076['E24'].border = Border(bottom=medium)
        RF1076['E29'].border = Border(bottom=medium)
        RF1076['E63'].border = Border(bottom= medium)

        RF1076['F9'].border = Border(top=medium, bottom=thin)
        RF1076['F10'].border = Border(bottom=thin)
        RF1076['F13'].border = Border(bottom=medium)
        RF1076['F14'].border = Border(bottom=medium, right=thin)
        RF1076['F19'].border = Border(bottom=medium)
        RF1076['F24'].border = Border(bottom=medium)
        RF1076['F29'].border = Border(bottom=medium)
        RF1076['F63'].border = Border(bottom= medium)
        
        RF1076['G9'].border = Border(top=medium, right=medium, bottom=thin)
        RF1076['G10'].border = Border(right=medium, bottom=thin)
        RF1076['G11'].border = Border(right=medium)
        RF1076['G12'].border = Border(right=medium)
        RF1076['G13'].border = Border(right=medium, bottom=medium)
        RF1076['G14'].border = Border(right=medium, bottom=medium)
        RF1076['G15'].border = Border(right=medium)
        RF1076['G16'].border = Border(right=medium)
        RF1076['G17'].border = Border(right=medium)
        RF1076['G18'].border = Border(right=medium)
        RF1076['G19'].border = Border(right=medium, bottom=medium)
        RF1076['G20'].border = Border(right=medium)
        RF1076['G21'].border = Border(right=medium)
        RF1076['G22'].border = Border(right=medium)
        RF1076['G23'].border = Border(right=medium)
        RF1076['G24'].border = Border(right=medium, bottom=medium)
        RF1076['G25'].border = Border(right=medium)
        RF1076['G26'].border = Border(right=medium)
        RF1076['G27'].border = Border(right=medium)
        RF1076['G28'].border = Border(right=medium)
        RF1076['G29'].border = Border(right=medium, bottom=medium)
        RF1076['G30'].border = Border(right=medium)
        RF1076['G31'].border = Border(right=medium)
        RF1076['G32'].border = Border(right=medium)
        RF1076['G33'].border = Border(right=medium)
        RF1076['G34'].border = Border(right=medium)
        RF1076['G35'].border = Border(right=medium)
        RF1076['G36'].border = Border(right=medium)
        RF1076['G37'].border = Border(right=medium)
        RF1076['G38'].border = Border(right=medium)
        RF1076['G39'].border = Border(right=medium)
        RF1076['G40'].border = Border(right=medium)
        RF1076['G41'].border = Border(right=medium)
        RF1076['G42'].border = Border(right=medium)
        RF1076['G43'].border = Border(right=medium)
        RF1076['G44'].border = Border(right=medium)
        RF1076['G45'].border = Border(right=medium)
        RF1076['G46'].border = Border(right=medium)
        RF1076['G47'].border = Border(right=medium)
        RF1076['G48'].border = Border(right=medium)
        RF1076['G49'].border = Border(right=medium)
        RF1076['G50'].border = Border(right=medium)
        RF1076['G51'].border = Border(right=medium)
        RF1076['G52'].border = Border(right=medium)
        RF1076['G53'].border = Border(right=medium)
        RF1076['G54'].border = Border(right=medium)
        RF1076['G55'].border = Border(right=medium)
        RF1076['G56'].border = Border(right=medium)
        RF1076['G57'].border = Border(right=medium)
        RF1076['G58'].border = Border(right=medium)
        RF1076['G59'].border = Border(right=medium)
        RF1076['G60'].border = Border(right=medium)
        RF1076['G61'].border = Border(right=medium)
        RF1076['G62'].border = Border(right=medium)
        RF1076['G63'].border = Border(right=medium, bottom= medium)
        
#Default values        
        RF1076['C9'].value='Product:'
        RF1076['C10'].value='Lot:'
        RF1076['C11'].value='Product Number:'
        RF1076['C12'].value='Container Tested:'
        RF1076['C13'].value='Container ID:'
        RF1076['F10'].value='Note:'
        
        RF1076['C14'].value='Parameter:'
        RF1076['D14'].value='Unit:'
        RF1076['F14'].value='Result:'
        RF1076['G14'].value='SD:'
        
        RF1076['B15'].value='Analyst:'
        RF1076['C15'].value='Particles (0.2 µm)'
        RF1076['D15'].value='Counts/mL'
        
        RF1076['C16'].value='Particles (0.3 µm)'
        RF1076['D16'].value='Counts/mL'
        
        RF1076['B17'].value='Date Analyzed:'
        RF1076['C17'].value='Particles (0.5 µm)'
        RF1076['D17'].value='Counts/mL'
        
        RF1076['C18'].value='Particles (1.0 µm)'
        RF1076['D18'].value='Counts/mL'
        
        RF1076['C19'].value='Particles (2.0 µm)'
        RF1076['D19'].value='Counts/mL'
   
        RF1076['B20'].value='Analyst:'
        RF1076['C20'].value='Assay (C1)'
        RF1076['D20'].value='% Wt'
        
        RF1076['C21'].value='Assay (C2)'
        RF1076['D21'].value='% Wt'
        
        RF1076['B22'].value='Date Analyzed:'
        RF1076['C22'].value='Assay (C3)'
        RF1076['D22'].value='% Wt'
        
        RF1076['C23'].value='Assay (C4)'
        RF1076['D23'].value='% Wt'
        
        RF1076['C24'].value='Assay (C5)'
        RF1076['D24'].value='% Wt'

        RF1076['B25'].value='Analyst:'
        RF1076['C25'].value='pH'
        RF1076['D25'].value='@'
        
        RF1076['C26'].value='pH 1% sol.'
        RF1076['D26'].value='@'
        
        RF1076['B27'].value='Date Analyzed:'
        RF1076['C27'].value='Specific Gravity'
        RF1076['D27'].value='@'
        
        RF1076['C28'].value='Surface Tension'
        RF1076['D28'].value='dynes/cm'
        
        RF1076['C29'].value='Surfactant Addition'
        RF1076['D29'].value='g/G'
               
        RF1076['B30'].value='Analyst:'
        RF1076['C30'].value='Aluminum (Al)'
        RF1076['D30'].value='ppb'
        RF1076['C31'].value='Antimony (Sb)'
        RF1076['D31'].value='ppb'
        RF1076['B32'].value='Date Analyzed:'
        RF1076['C32'].value='Arsenic (As)'
        RF1076['D32'].value='ppb'
        RF1076['C33'].value='Barium (Ba)'
        RF1076['D33'].value='ppb'
        RF1076['C34'].value='Beryllium (Be)'
        RF1076['D34'].value='ppb'
        RF1076['C35'].value='Bismuth (Bi)'
        RF1076['D35'].value='ppb'
        RF1076['C36'].value='Boron (B)'
        RF1076['D36'].value='ppb'
        RF1076['C37'].value='Cadmium (Cd)'
        RF1076['D37'].value='ppb'
        RF1076['C38'].value='Calcium (Ca)'
        RF1076['D38'].value='ppb'
        RF1076['C39'].value='Chromium (Cr)'
        RF1076['D39'].value='ppb'
        RF1076['C40'].value='Cobalt (Co)'
        RF1076['D40'].value='ppb'
        RF1076['C41'].value='Copper (Cu)'
        RF1076['D41'].value='ppb'
        RF1076['C42'].value='Gallium (Ga)'
        RF1076['D42'].value='ppb'
        RF1076['C43'].value='Germanium (Ge)'
        RF1076['D43'].value='ppb'
        RF1076['C44'].value='Gold (Au)'
        RF1076['D44'].value='ppb'
        RF1076['C45'].value='Iron (Fe)'
        RF1076['D45'].value='ppb'
        RF1076['C46'].value='Lead (Pb)'
        RF1076['D46'].value='ppb'
        RF1076['C47'].value='Lithium (Li)'
        RF1076['D47'].value='ppb'
        RF1076['C48'].value='Magnesium (Mg)'
        RF1076['D48'].value='ppb'
        RF1076['C49'].value='Manganese (Mn)'
        RF1076['D49'].value='ppb'
        RF1076['C50'].value='Molybdenum (Mo)'
        RF1076['D50'].value='ppb'
        RF1076['C51'].value='Nickle (Ni)'
        RF1076['D51'].value='ppb'
        RF1076['C52'].value='Niobium (Nb)'
        RF1076['D52'].value='ppb'
        RF1076['C53'].value='Potassium (K)'
        RF1076['D53'].value='ppb'
        RF1076['C54'].value='Silver (Ag)'
        RF1076['D54'].value='ppb'
        RF1076['C55'].value='Sodium (Na)'
        RF1076['D55'].value='ppb'
        RF1076['C56'].value='Strontium (Sr)'
        RF1076['D56'].value='ppb'
        RF1076['C57'].value='Tantalum (Ta)'
        RF1076['D57'].value='ppb'
        RF1076['C58'].value='Thallium (Tl)'
        RF1076['D58'].value='ppb'
        RF1076['C59'].value='Tin (Sn)'
        RF1076['D59'].value='ppb'
        RF1076['C60'].value='Titanium (Ti)'
        RF1076['D60'].value='ppb'
        RF1076['C61'].value='Vanadium (V)'
        RF1076['D61'].value='ppb'
        RF1076['C62'].value='Zinc (Zn)'
        RF1076['D62'].value='ppb'
        RF1076['C63'].value='Zirconium (Zr)'
        RF1076['D63'].value='ppb'
        
#Actual Values
#Product ID values
        RF1076['D10'].value = str(zuple[1][:])
        try:
            RF1076['D11'].value = round(zuple[2], 0)
        except TypeError:
            RF1076['D11'].value = zuple[2]
            
        RF1076['D12'].value = str(zuple[12][:])
        RF1076['D13'].value = str(zuple[6])
#Particle Data
        if zuple[0] == 'X' or zuple[0] == 'x':
            RF1076['B16'].value= str(zuple[4][:])
            RF1076['B18'].value= zuple[5].strftime("%m/%d/%Y")
            
            RF1076['F15'].value= zuple[7]
            RF1076['F16'].value= zuple[8]
            RF1076['F17'].value= zuple[9]
            RF1076['F18'].value= zuple[10]
            RF1076['F19'].value= zuple[11]
        
#Assay Data        
        if zuple[13] == 'X' or zuple[13] == 'x':
            RF1076['B21'].value= str(zuple[14][:])
            RF1076['B23'].value= zuple[15].strftime("%m/%d/%Y")
            
            SUB = str.maketrans("0123456789", "₀₁₂₃₄₅₆₇₈₉")
            try1=str(zuple[16]).translate(SUB)
            try2=str(zuple[17]).translate(SUB)
            try3=str(zuple[18]).translate(SUB)
            
            RF1076['E20'].value= try1
            RF1076['F20'].value= round(zuple[19], 2)
            RF1076['G20'].value= '± ' + str(round(zuple[20], 2))
            
            if zuple[21] >= 0:
                RF1076['F21'].value= round(zuple[21], 2)
                RF1076['G21'].value= '± ' + str(round(zuple[22], 2))
                RF1076['E21'].value= try2
                
            else:
                RF1076['F21'].value= ''
                RF1076['G21'].value= ''
                RF1076['E21'].value= ''
                
            if zuple[23] >= 0:
                RF1076['F22'].value= round(zuple[23], 2)
                RF1076['G22'].value= '± ' + str(round(zuple[24], 2))
                RF1076['E22'].value= try3
            
            else:
                RF1076['F22'].value= ''
                RF1076['G22'].value= ''
                RF1076['E22'].value= ''                
            
            if zuple[17] == 'BOR:' or zuple[17] == 'First Bottle:':
                RF1076['E21'].value= try2
            
        if zuple[25] == 'X' or zuple[25] == 'x':
            RF1076['B26'].value= str(zuple[26][:])
            RF1076['B28'].value= zuple[27].strftime("%m/%d/%Y")
            
            if zuple[28] == 'ST':
                RF1076['F28'].value= zuple[32]
            if zuple[28] == 'SG':
                RF1076['F27'].value= zuple[31]
            if zuple[28] == '1pH':
                RF1076['F26'].value= zuple[30]
            if zuple[28] == 'pH':
                RF1076['F25'].value= zuple[29]

#Metals Data
        if zuple[33] == 'X' or zuple[33] == 'x':
            phosmdl=[3,3,3,3,4,3,5,2,3,3,2,2,2,2,1,1,2,2,3,2,2,2,1,2,2,3,2,2,2,2,6,2,3,2]
            if zuple[36] == 'PHOS-SA' or zuple[36] == 'PHOS34-SA':
                RF1076['E30'].value= phosmdl[0]   
                RF1076['E31'].value= phosmdl[1]
                RF1076['E32'].value= phosmdl[2]
                RF1076['E33'].value= phosmdl[3]
                RF1076['E34'].value= phosmdl[4]
                RF1076['E35'].value= phosmdl[5]
                RF1076['E36'].value= phosmdl[6]
                RF1076['E37'].value= phosmdl[7]
                RF1076['E38'].value= phosmdl[8]
                RF1076['E39'].value= phosmdl[9]
                RF1076['E40'].value= phosmdl[10]
                RF1076['E41'].value= phosmdl[11]
                RF1076['E42'].value= phosmdl[12]
                RF1076['E43'].value= phosmdl[13]
                RF1076['E44'].value= phosmdl[14]
                RF1076['E45'].value= phosmdl[15]
                RF1076['E46'].value= phosmdl[16]
                RF1076['E47'].value= phosmdl[17]
                RF1076['E48'].value= phosmdl[18]
                RF1076['E49'].value= phosmdl[19]
                RF1076['E50'].value= phosmdl[20]
                RF1076['E51'].value= phosmdl[21]
                RF1076['E52'].value= phosmdl[22]
                RF1076['E53'].value= phosmdl[23]
                RF1076['E54'].value= phosmdl[24]
                RF1076['E55'].value= phosmdl[25]
                RF1076['E56'].value= phosmdl[26]
                RF1076['E57'].value= phosmdl[27]
                RF1076['E58'].value= phosmdl[28]
                RF1076['E59'].value= phosmdl[29]
                RF1076['E60'].value= phosmdl[30]
                RF1076['E61'].value= phosmdl[31]
                RF1076['E62'].value= phosmdl[32]
                RF1076['E63'].value= phosmdl[33]

                
            RF1076['B31'].value= str(zuple[34][:])
            RF1076['B33'].value= zuple[35].strftime("%m/%d/%Y") 
            
            metals = []
            
            for x in zuple[37:71]:
                if type(x)==str:
                    metals.append(x)
                else:
                    metals.append(round(x, 1))
            
                
            RF1076['F30'].value= metals[0]   
            RF1076['F31'].value= metals[1]
            RF1076['F32'].value= metals[2]
            RF1076['F33'].value= metals[3]
            RF1076['F34'].value= metals[4]
            RF1076['F35'].value= metals[5]
            RF1076['F36'].value= metals[6]
            RF1076['F37'].value= metals[7]
            RF1076['F38'].value= metals[8]
            RF1076['F39'].value= metals[9]
            RF1076['F40'].value= metals[10]
            RF1076['F41'].value= metals[11]
            RF1076['F42'].value= metals[12]
            RF1076['F43'].value= metals[13]
            RF1076['F44'].value= metals[14]
            RF1076['F45'].value= metals[15]
            RF1076['F46'].value= metals[16]
            RF1076['F47'].value= metals[17]
            RF1076['F48'].value= metals[18]
            RF1076['F49'].value= metals[19]
            RF1076['F50'].value= metals[20]
            RF1076['F51'].value= metals[21]
            RF1076['F52'].value= metals[22]
            RF1076['F53'].value= metals[23]
            RF1076['F54'].value= metals[24]
            RF1076['F55'].value= metals[25]
            RF1076['F56'].value= metals[26]
            RF1076['F57'].value= metals[27]
            RF1076['F58'].value= metals[28]
            RF1076['F59'].value= metals[29]
            RF1076['F60'].value= metals[30]
            RF1076['F61'].value= metals[31]
            RF1076['F62'].value= metals[32]
            RF1076['F63'].value= metals[33]

#Center Values/Banket formatting
        for col in RF1076.columns:
            for cell in col:
                alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                cell.font = font1
                cell.alignment = alignment_obj

#Special Formatting & Values (actual and default) Must keep after centering/blanket formating group.             
        RF1076.add_image(img, 'A1')
        
        b=RF1076['G1']
        b.value='RF 1076'
        b.font = Font(name='Times New Roman', size=24, bold=False, italic=False,
                        underline='none',strike=False,color='000000')

        c=RF1076['A2']
        c.value='HPP Lab Data Report'
        c.font = Font(name='Times New Roman', size=24, bold=False, italic=False,
                               underline='none',strike=False,color='000000')

        d=RF1076['G2']
        d.value='Rev 2'
        d.font = Font(name='Times New Roman', size=24, bold=False, italic=False,
                              underline='none',strike=False,color='000000')
        e=RF1076['A3']
        e.value='Reviewed and Approved By Daniel Williams'
        e.font = Font(name='Times New Roman', size=11, bold=True, italic=False,
                               underline='none',strike=False,color='000000') 

        f=RF1076['G3']
        f.value='Date: 08.12.20'
        f.font = Font(name='Times New Roman', size=14, bold=False, italic=False,
                               underline='none',strike=False,color='000000') 

        g=RF1076['A4']
        g.value='This document is uncontrolled in a printed or electronically transmitted version unless clearly identified as controlled.'
        g.font = Font(name='Arial', size=8, bold=False, italic=False,
                              underline='none',strike=False,color='000000')
                   
        RF1076['B6'].value='Report Created by:'
        RF1076['B6'].alignment=Alignment(horizontal='right')
        RF1076['B7'].value='Date:'
        RF1076['B7'].alignment=Alignment(horizontal='right')
        RF1076['C6'].value='L. Marx'
        RF1076['C7'].value=date.today().strftime("%m/%d/%Y")
        RF1076['F7'].value='=IF(F6<>"", IF(F7="",TEXT(TODAY(),"mm/dd/yyyy"),F7), "")'
        RF1076['E6'].value='Reviewed by:'
        RF1076['E6'].alignment=Alignment(horizontal='right')
        RF1076['E7'].value='Date:'
        RF1076['E7'].alignment=Alignment(horizontal='right')
        
        RF1076['D9'].value = str(zuple[3][:])
        RF1076['D9'].alignment=Alignment(horizontal='left')
        
#Export and save report                
        try:
            Path3 = os.path.join(Path2, str(zuple[1][:]) + ' ' + str(zuple[3][:]) + '.xlsx')
            wb.save(Path3)
        except PermissionError:
            sg.popup('Close old reports and try generating new reports again.')
            return
        
#LindseyLot Generator    
    completeName = os.path.join(Path2, 'Lot #s to Email.txt')
    dflot = str(df.Lot.unique())
    dflot1 = dflot.replace("' ", ' \n')
    dflot2 = dflot1.replace("'", '')
    dflot3 = dflot2.replace("[", '')
    dflot4 = dflot3.replace("]", '')
    dflot5 = dflot4.replace(" ", '')
    file1 = open(completeName, "w") 
    file1.write(dflot5)
    file1.close()
    
    sg.popup('Success')
    
def counter(y1, z1):
#build DF
    data = pd.ExcelFile(workingpath)
    dfa = pd.DataFrame(pd.read_excel(data, 'InLabLogdata'))
    
    dfaa = dfa[['SDT', 'SA', 'SO', 'SM', 'SCT', 'SR']]
    
    dfab = dfa[['PDT', 'P02um', 'RP']]
    
    dfac = dfa[['RA', 'ADT', 'A1', 'A2', 'A3', 'AC2']]
    
    dfad = dfa[['RO', 'ODT', 'Oanalysis', 'OpH', 'O1pH', 'OSG', 'OST']]
    
    dfae = dfa[['RM', 'MDT', 'Mmethod']]
#convert dates and filter   
    try:
        dfaa['SDT']= pd.to_datetime(dfaa['SDT'])
        dfaa['SDT']= dfaa['SDT'].dt.date
        dfaa1= dfaa[(dfaa['SDT'] >= y1)]
        dfaa2= dfaa1[(dfaa1['SDT'] <= z1)]
    except:
        sg.popup('Check Sampling Date Formating in "Lab Data" sheet.')
        return
   
    try:
        dfab['PDT']= pd.to_datetime(dfab['PDT'])
        dfab['PDT']= dfab['PDT'].dt.date
        dfab1= dfab[(dfab['PDT'] >= y1)]
        dfab2= dfab1[(dfab1['PDT'] <= z1)]
    except:
        sg.popup('Check Particle Date Formating in "Lab Data" sheet.')
        return
    
    try:
        dfac['ADT']= pd.to_datetime(dfac['ADT'])
        dfac['ADT']= dfac['ADT'].dt.date
        dfac1= dfac[(dfac['ADT'] >= y1)]
        dfac2= dfac1[(dfac1['ADT'] <= z1)]
    except:
        sg.popup('Check Assay Date Formating in "Lab Data" sheet.')
        return
    
    try:
        dfad['ODT']= pd.to_datetime(dfad['ODT'])
        dfad['ODT']= dfad['ODT'].dt.date
        dfad1= dfad[(dfad['ODT'] >= y1)]
        dfad2= dfad1[(dfad1['ODT'] <= z1)]
    except:
        sg.popup('Check "Other" Date Formating in "Lab Data" sheet.')
        return
    
    try:
        dfae['MDT']= pd.to_datetime(dfae['MDT'])
        dfae['MDT']= dfae['MDT'].dt.date
        dfae1= dfae[(dfae['MDT'] >= y1)]
        dfae2= dfae1[(dfae1['MDT'] <= z1)]
    except:
        sg.popup('Check Metals Date Formating in "Lab Data" sheet.')
        return
    
#Bottles and samples
    m1=dfaa2['SM'].sum()
    o1=dfaa2['SO'].sum()
    as1=dfaa2['SA'].sum()
    ct1=dfaa2['SCT'].sum()
    r1=dfaa2['SR'].sum()
#Particles
    dfab2.dropna(subset = ['P02um'], inplace=True)
    p2= dfab2['P02um'].count()
    Parts3 = dfab2[dfab2['RP']=='X']
    p1= Parts3['P02um'].count()
    
#setup for assays    
    dfac2.dropna(subset = ['A1'], inplace=True)
    hf100to1a = dfac2[dfac2['AC2']== 'First Bottle:']
    hf100to1a.dropna(subset = ['A3'], inplace = True)
    hf1001fb = hf100to1a['AC2'].count()
    hf100to1aa = dfac2[dfac2['AC2']== 'BOR:']
    hf100to1aa.dropna(subset = ['A3'], inplace = True)
    hf1001bor = hf100to1aa['AC2'].count()
    hf1001 = hf1001fb+hf1001bor
    
#setup for reported Assays
    Assay3 = dfac2[dfac2['RA']=='X']
    hf100to1ax = Assay3[Assay3['AC2']== 'First Bottle:']
    hf1001fbx = hf100to1ax['AC2'].count()
    hf100to1aax = Assay3[Assay3['AC2']== 'BOR:']
    hf1001borx = hf100to1aax['AC2'].count()
    hf1001x = hf1001fbx+hf1001borx

#All assays count
    all_assays= dfac2['A1'].count()
    min_2 = dfac2['A2'].count()
    all_three= dfac2['A3'].count()-hf1001
    
#reported assay count    
    all_assaysX= Assay3['A1'].count()
    min_2X= Assay3['A2'].count()
    all_threeX= Assay3['A3'].count()-hf1001x
    
#Assay compiled    
    a= all_assaysX-min_2X
    aa= min_2X-all_threeX
    aaa= all_threeX
    aaa2= all_three
    aa2= min_2-all_three
    a2= all_assays-min_2
#Other counter
    dfad2.dropna(subset = ['Oanalysis'], inplace=True)
    dfad2ST = dfad2[dfad2['Oanalysis']=='ST']
    dfad2SG = dfad2[dfad2['Oanalysis']=='SG']
    ST= dfad2ST['Oanalysis'].count()
    SG= dfad2SG['Oanalysis'].count()
#Metals counter
    dfae2.dropna(subset = ['Mmethod'], inplace=True)
    dfae2PHOS1 = dfae2[dfae2['Mmethod']=='PHOS34-SA']
    Phosa = dfae2PHOS1['Mmethod'].count()
    dfae2PHOS2 = dfae2[dfae2['Mmethod']=='PHOS-SA']
    Phosb = dfae2PHOS2['Mmethod'].count()
   #dfae2Other = dfae2[dfae2['Mmethod']!= 'PHOS34-SA' or dfae2['Mmethod']!= 'PHOS-SA']
    Phos = Phosa + Phosb
   #otherm = dfae2Other['Mmethod'].count()

#Counter report maker
    if bb == 'Yes':
        wb = openpyxl.Workbook() 
        bob = wb.active 
        bob.title = "counter"
        
        bob.column_dimensions['A'].width = 19
        bob.column_dimensions['B'].width = 16.75
        bob.column_dimensions['C'].width = 17
        bob.column_dimensions['D'].width = 15
        bob.column_dimensions['E'].width = 16
        
        font1= Font(name='Calibri', size=11, 
                        bold=False, italic=False,
                        underline='none',strike=False,color='000000')
        thin = Side(border_style="thin", color="000000")
                    
        bob['A1'].value ='Analytical All:'
        bob['B1'].value ='Particles Analyzed'
        bob['C1'].value ='Assays Analyzed'
        bob['D1'].value ='Metals Analyzed'
        bob['E1'].value ='Other Analyses'
        
        bob['A3'].value ='Assay Breakdown:'
        bob['B3'].value ='One Analyte'
        bob['C3'].value ='Two Analyte'
        bob['D3'].value ='Three Analyte'
        
        bob['A5'].value ='Analytical Reported:'
        bob['B5'].value ='Particles Analyzed'
        bob['C5'].value ='Assays Analyzed'
        bob['D5'].value ='Metals Analyzed'
        bob['E5'].value ='Other Analyses'
        
        bob['A7'].value ='Assay Breakdown:'
        bob['B7'].value ='One Analyte'
        bob['C7'].value ='Two Analyte'
        bob['D7'].value ='Three Analyte'
        
        bob['A9'].value ='Sampling Summary:'
        bob['B9'].value ='Total Samples'
        bob['C9'].value ='In Hosuse Samples'
        bob['D9'].value ='CT Samples'
        bob['E9'].value ='Retain'
        
        bob['A11'].value ='Bottle Consumption:'
        bob['B11'].value ='500 mL Bottles'
        bob['C11'].value ='125 mL Bottles'
        bob['D11'].value ='Teflon Bottles'
        
        bob['B2'].value = p2
        bob['C2'].value = all_assays
        bob['D2'].value = Phos
        bob['E2'].value = SG+ST
        
        bob['B6'].value = p1
        bob['C6'].value = all_assaysX
        bob['D6'].value = Phos
        bob['E6'].value = SG+ST

        bob['B4'].value = a2
        bob['C4'].value = aa2
        bob['D4'].value = aaa2
        
        bob['B8'].value = a
        bob['C8'].value = aa
        bob['D8'].value = aaa
        
        bob['B10'].value = m1+o1+as1+ct1+r1
        bob['C10'].value = as1+o1+m1
        bob['D10'].value = ct1
        bob['E10'].value = r1
    
        bob['B12'].value = ct1+r1
        bob['C12'].value = o1+as1
        bob['D12'].value = m1
        
        for col in bob.columns:
                for cell in col:
                    alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                    cell.font = font1
                    cell.alignment = alignment_obj
                    cell.border = Border(top=thin, left=thin, bottom=thin, right=thin)
                    
        Path3 = os.path.join(path, 'Counter Output ' + str(y1) + '-' + str(z1) + '.xlsx')
        wb.save(Path3)
        
#Generic Output maker    
    results= {
             '': ['Reported', 'All'],
             'Total Samples Drawn:': ['n/a', m1+o1+as1+ct1+r1],
             'Metals Samples': ['n/a', m1], 'Other Samples': ['n/a', o1],'Assay Samples': ['n/a', as1],
             'CT Samples': ['n/a', ct1],'Retain Samples': ['n/a', r1],
             '500 mL Bottles': ['n/a', ct1+r1], '125 mL Bottles': ['n/a', o1+as1], 'Teflon Bottles': ['n/a', m1],
             'Particles': [p1,p2], 'Assays': [all_assaysX, all_assays], 'Metals': [Phos, 'n/a'],
             'Surface Tension': [ST, 'n/a'], 'Specific Gravity': [SG, 'n/a'],
             'One Component Assays': [a,a2], 'Two Component Assays': [aa,aa2], 
             'Three Component Assays': [aaa,aaa2]
              }
    
    dfan = pd.DataFrame(data=results)
    df = dfan.transpose()
    
    printme = tabulate(df[1:], headers=[str(Date1a) + ' - ' + str(Date2a), 'Reported', 'All'])
    sg.popup(printme)
                    
def pullparts(PN, sort, report, datex): 
    
    data = pd.ExcelFile(workingpath)
    dfa = pd.DataFrame(pd.read_excel(data, 'InLabLogdata'))
    
    dfbc = pd.DataFrame(pd.read_excel(data,  'Spec'))
    dfb = dfbc[dfbc['Wking']=='X']
    
    dfaa = pd.merge(dfa, dfb, on = 'PN', how = "left")
    dfaa['PDT']= pd.to_datetime(dfaa['PDT'])
    dfaa['PDT']= dfaa['PDT'].dt.date
    
    plotdf= dfaa[(dfaa['PDT'] >= datex)]
    plotdf= plotdf[(plotdf['PN']==PN)]
    
    if sort=='Lot':
        plotdf.sort_values('Lot', inplace = True)
                
    elif sort=='Date':
        plotdf.sort_values('PDT', inplace = True)
                
    
    if report=='Reported':
        is_reported = plotdf['RP']=='X'
        plotdf1 = plotdf[is_reported]
        specialname = 'Reported Data'
    elif report =='All':
        plotdf1 = plotdf
        specialname = 'All Data'
        
    prodname = str(plotdf1.iloc[0]['Product_x'])
    #dfaa.to_excel(r'C:\Users\HPP Assay\Desktop\Dump\dfa.xlsx', sheet_name='1')
    #plotdf1.to_excel(r'C:\Users\HPP Assay\Desktop\Dump\plotdf1.xlsx', sheet_name='1')

#0.2 plot    
    if all(isinstance(item, int) for item in plotdf1.PSp02) == False:
        if all(isinstance(item, float) for item in plotdf1.PSp02) == False:
            pass
    else:
        w = 20
        h = 5
        d = 70
        plt.figure(figsize=(w, h), dpi=d)
        plt.xticks(rotation=75)
        plt.tick_params(axis='x', labelsize=10)
        plt.tick_params(bottom=True, top=False, left=True, right=True)
        plt.ylabel('Counts per mL')
        plt.xlabel('Lot (Date Range: ' + str(datex.strftime('%b %d,%y')) + ' - ' + str(datetime.date.today().strftime('%b %d,%y')) + ')')
        plt.title('Manufacturing Trend at 0.2 um for ' + prodname + ' (' + str(PN) + ')')
        plt.ylim(0, max(plotdf1.P02um)+100)
        red_patch = mpatches.Patch(color='red', label='Control Limit +/- 2SD')
        blue_patch = mpatches.Patch(color='blue', label='Product Data')
        yellow_patch = mpatches.Patch(color='orange', label='Product Spec: ' + str(statistics.mean(plotdf1.PSp02)))
        purple_patch = mpatches.Patch(color='purple', label='Average Particle Count: ' + str(round(statistics.mean(plotdf1.P02um), 0)))
        white3 = mpatches.Patch(color='white', label='Filters: ' + str(specialname) + ' sorted by ' + str(sort))
        plt.legend(handles=[white3, red_patch, blue_patch, yellow_patch, purple_patch])
        plt.axhline(y=statistics.mean(plotdf1.P02um), color='purple', linestyle='--')
    
        x = plotdf1.Lot
    
        y1 = plotdf1.P02um
                
        plt.axhline(y=statistics.mean(plotdf1.PSp02), color='orange', linestyle='-')
        plt.axhline(y=statistics.mean(plotdf1.P02um)+(2*(statistics.stdev(plotdf1.P02um))), color='r', linestyle='-')
        plt.axhline(y=statistics.mean(plotdf1.P02um)-(2*(statistics.stdev(plotdf1.P02um))), color='r', linestyle='-') 
        plt.plot(x, y1, marker='o')
        path02 = os.path.join(pathpart, prodname + ' 0.2 um ' + str(datex.strftime('%b %d,%y')) + '_' + str(datetime.date.today().strftime('%b %d,%y')) + '.png')
        plt.savefig(path02, bbox_inches='tight')
        
#0.3 plot 
    if all(isinstance(item, int) for item in plotdf1.PSp03) == False:
        if all(isinstance(item, float) for item in plotdf1.PSp03) == False:
            pass
    
    else:
        w = 20
        h = 5
        d = 70
        plt.figure(figsize=(w, h), dpi=d)
        plt.xticks(rotation=75)
        plt.tick_params(axis='x', labelsize=10)
        plt.tick_params(bottom=True, top=False, left=True, right=True)
        plt.ylabel('Counts per mL')
        plt.xlabel('Lot (Date Range: ' + str(datex.strftime('%b %d,%y')) + ' - ' + str(datetime.date.today().strftime('%b %d,%y')) + ')')
        plt.title('Manufacturing Trend at 0.3 um for ' + prodname + ' (' + str(PN) + ')')
        plt.ylim(0, max(plotdf1.P03um)+100)
        red_patch = mpatches.Patch(color='red', label='Control Limit +/- 2SD')
        blue_patch = mpatches.Patch(color='blue', label='Product Data')
        yellow_patch = mpatches.Patch(color='orange', label='Product Spec: ' + str(statistics.mean(plotdf1.PSp03)))
        purple_patch = mpatches.Patch(color='purple', label='Average Particle Count: ' + str(round(statistics.mean(plotdf1.P03um), 0)))
        white3 = mpatches.Patch(color='white', label='Filters: ' + str(specialname) + ' sorted by ' + str(sort))
        plt.legend(handles=[white3, red_patch, blue_patch, yellow_patch, purple_patch])
        plt.axhline(y=statistics.mean(plotdf1.P03um), color='purple', linestyle='--')
    
        x = plotdf1.Lot
    
        y1 = plotdf1.P03um
               
        plt.axhline(y=statistics.mean(plotdf1.PSp03), color='orange', linestyle='-')
        plt.axhline(y=statistics.mean(plotdf1.P03um)+(2*(statistics.stdev(plotdf1.P03um))), color='r', linestyle='-')
        plt.axhline(y=statistics.mean(plotdf1.P03um)-(2*(statistics.stdev(plotdf1.P03um))), color='r', linestyle='-')
        plt.plot(x, y1, marker='o')
        path03 = os.path.join(pathpart, prodname + ' 0.3 um ' + str(datex.strftime('%b %d,%y')) + '_' + str(datetime.date.today().strftime('%b %d,%y')) + '.png')
        plt.savefig(path03, bbox_inches='tight')
        
#0.5 plot 
    if all(isinstance(item, int) for item in plotdf1.PSp05) == False:
        if all(isinstance(item, float) for item in plotdf1.PSp05) == False:
            pass
    else:
        w = 20
        h = 5
        d = 70
        plt.figure(figsize=(w, h), dpi=d)
        plt.xticks(rotation=75)
        plt.tick_params(axis='x', labelsize=10)
        plt.tick_params(bottom=True, top=False, left=True, right=True)
        plt.ylabel('Counts per mL')
        plt.xlabel('Lot (Date Range: ' + str(datex.strftime('%b %d,%y')) + ' - ' + str(datetime.date.today().strftime('%b %d,%y')) + ')')
        plt.title('Manufacturing Trend at 0.5 um for ' + prodname + ' (' + str(PN) + ')')
        plt.ylim(0, max(plotdf1.P05um)+50)
        red_patch = mpatches.Patch(color='red', label='Control Limit +/- 2SD')
        blue_patch = mpatches.Patch(color='blue', label='Product Data')
        yellow_patch = mpatches.Patch(color='orange', label='Product Spec: ' + str(statistics.mean(plotdf1.PSp05)))
        purple_patch = mpatches.Patch(color='purple', label='Average Particle Count: ' + str(round(statistics.mean(plotdf1.P05um), 0)))
        white3 = mpatches.Patch(color='white', label='Filters: ' + str(specialname) + ' sorted by ' + str(sort))
        plt.legend(handles=[white3, red_patch, blue_patch, yellow_patch, purple_patch])
        plt.axhline(y=statistics.mean(plotdf1.P05um), color='purple', linestyle='--')
    
        x = plotdf1.Lot
    
        y1 = plotdf1.P05um
        
        plt.axhline(y=statistics.mean(plotdf1.PSp05), color='orange', linestyle='-')
        plt.axhline(y=statistics.mean(plotdf1.P05um)+(2*(statistics.stdev(plotdf1.P05um))), color='r', linestyle='-')
        plt.axhline(y=statistics.mean(plotdf1.P05um)-(2*(statistics.stdev(plotdf1.P05um))), color='r', linestyle='-')
        plt.plot(x, y1, marker='o')
        path05 = os.path.join(pathpart, prodname + ' 0.5 um ' + str(datex.strftime('%b %d,%y')) + '_' + str(datetime.date.today().strftime('%b %d,%y')) + '.png')
        plt.savefig(path05, bbox_inches='tight')
            
#1.0 plot 
    if all(isinstance(item, int) for item in plotdf1.PSp10) == False:
        if all(isinstance(item, float) for item in plotdf1.PSp10) == False:
            pass
    
    else:
        w = 20
        h = 5
        d = 70
        plt.figure(figsize=(w, h), dpi=d)
        plt.xticks(rotation=75)
        plt.tick_params(axis='x', labelsize=10)
        plt.tick_params(bottom=True, top=False, left=True, right=True)
        plt.ylabel('Counts per mL')
        plt.xlabel('Lot (Date Range: ' + str(datex.strftime('%b %d,%y')) + ' - ' + str(datetime.date.today().strftime('%b %d,%y')) + ')')
        plt.title('Manufacturing Trend at 1.0 um for ' + prodname + ' (' + str(PN) + ')')
        plt.ylim(0, max(plotdf1.P10um)+10)
        red_patch = mpatches.Patch(color='red', label='Control Limit +/- 2SD')
        blue_patch = mpatches.Patch(color='blue', label='Product Data')
        yellow_patch = mpatches.Patch(color='orange', label='Product Spec: ' + str(statistics.mean(plotdf1.PSp10)))
        purple_patch = mpatches.Patch(color='purple', label='Average Particle Count: ' + str(round(statistics.mean(plotdf1.P10um), 0)))
        white3 = mpatches.Patch(color='white', label='Filters: ' + str(specialname) + ' sorted by ' + str(sort))
        plt.legend(handles=[white3, red_patch, blue_patch, yellow_patch, purple_patch])
        plt.axhline(y=statistics.mean(plotdf1.P10um), color='purple', linestyle='--')
        
        x = plotdf1.Lot
    
        y1 = plotdf1.P10um
       
        plt.axhline(y=statistics.mean(plotdf1.PSp10), color='orange', linestyle='-')
        plt.axhline(y=statistics.mean(plotdf1.P10um)+(2*(statistics.stdev(plotdf1.P10um))), color='r', linestyle='-')
        plt.axhline(y=statistics.mean(plotdf1.P10um)-(2*(statistics.stdev(plotdf1.P10um))), color='r', linestyle='-')
        plt.plot(x, y1, marker='o')
        path10 = os.path.join(pathpart, prodname + ' 1.0 um ' + str(datex.strftime('%b %d,%y')) + '_' + str(datetime.date.today().strftime('%b %d,%y')) + '.png')
        plt.savefig(path10, bbox_inches='tight')
        
#2.0 plot 
    if all(isinstance(item, int) for item in plotdf1.PSp20) == False:
        if all(isinstance(item, float) for item in plotdf1.PSp20) == False:
            pass
            
    else:
        w = 20
        h = 5
        d = 70
        plt.figure(figsize=(w, h), dpi=d)
        plt.xticks(rotation=75)
        plt.tick_params(axis='x', labelsize=10)
        plt.tick_params(bottom=True, top=False, left=True, right=True)
        plt.ylabel('Counts per mL')
        plt.xlabel('Lot (Date Range: ' + str(datex.strftime('%b %d,%y')) + ' - ' + str(datetime.date.today().strftime('%b %d,%y')) + ')')
        plt.title('Manufacturing Trend at 2.0 um for ' + prodname + ' (' + str(PN) + ')')
        plt.ylim(0, max(plotdf1.P20um)+10)
        red_patch = mpatches.Patch(color='red', label='Control Limit +/- 2SD')
        blue_patch = mpatches.Patch(color='blue', label='Product Data')
        yellow_patch = mpatches.Patch(color='orange', label='Product Spec: ' + str(statistics.mean(plotdf1.PSp20)))
        purple_patch = mpatches.Patch(color='purple', label='Average Particle Count: ' + str(round(statistics.mean(plotdf1.P20um), 0)))
        white3 = mpatches.Patch(color='white', label='Filters: ' + str(specialname) + ' sorted by ' + str(sort))
        plt.legend(handles=[white3, red_patch, blue_patch, yellow_patch, purple_patch])
        plt.axhline(y=statistics.mean(plotdf1.P20um), color='purple', linestyle='--')
        
        
        x = plotdf1.Lot
    
        y1 = plotdf1.P20um
    
        plt.axhline(y=statistics.mean(plotdf1.PSp20), color='orange', linestyle='-')
        plt.axhline(y=statistics.mean(plotdf1.P20um)+(2*(statistics.stdev(plotdf1.P20um))), color='r', linestyle='-')
        plt.axhline(y=statistics.mean(plotdf1.P20um)-(2*(statistics.stdev(plotdf1.P20um))), color='r', linestyle='-')
        plt.plot(x, y1, marker='o')
        path20 = os.path.join(pathpart, prodname + ' 2.0 um ' + str(datex.strftime('%b %d,%y')) + '_' + str(datetime.date.today().strftime('%b %d,%y')) + '.png')
        plt.savefig(path20, bbox_inches='tight')
        
def pullassay(x1, sort, report, datex):
    data = pd.ExcelFile(workingpath)
    
    dfa = pd.DataFrame(pd.read_excel(data, 'InLabLogdata'))
    
    dfbc = pd.DataFrame(pd.read_excel(data,  'Spec'))
    dfb = dfbc[dfbc['Wking']=='X']
    
    dfaa = pd.merge(dfa, dfb, on = 'PN', how = "left")
    dfaa['ADT']= pd.to_datetime(dfaa['ADT'])
    dfaa['ADT']= dfaa['ADT'].dt.date
    
    plotdf= dfaa[(dfaa['ADT'] >= datex)]      
    
####
    if sort=='Lot':
        plotdf.sort_values('Lot', inplace = True)
        #plotdf.to_excel(r'C:\Users\HPP Assay\Desktop\Dump\plotdf.xlsx', sheet_name='1')     
    elif sort=='Date':
        plotdf.sort_values('ADT', inplace = True)
           
####              
    if type(x1)==str:
        plotdfa = plotdf[plotdf['Product_x']==x1]
    elif type(x1)==int:
        plotdfa = plotdf[plotdf['PN']==x1]
        #plotdfa.to_excel(r'C:\Users\HPP Assay\Desktop\Excel DUmp\plotdfa.xlsx', sheet_name='1')
    
####        
    if report =='Reported':
        is_reported = plotdfa['RA']=='X'
        plotdfa1 = plotdfa[is_reported]
        plotdfa1.reset_index(inplace = True)
        specialname = 'Reported Data'
        #plotdfa1.to_excel(r'C:\Users\HPP Assay\Desktop\Dump\plotdfa1.xlsx', sheet_name='1')
        #return
    
    elif report=='SPC':
        is_spc = plotdfa['RA']=='XQ'
        plotdfa1 = plotdfa[is_spc]
        plotdfa1.reset_index(inplace = True)
        specialname = 'SPC Data'
        #plotdfa1.to_excel(r'C:\Users\HPP Assay\Desktop\Excel DUmp\plotdfa1.xlsx', sheet_name='1')
        
    prodname = str(plotdfa1.iloc[0]['Product_x'])
    pna = round(plotdfa1.iloc[0]['PN'])
    PN = str(pna)
####plot A1   
    if statistics.mean(map(float, plotdfa1.A1)) == 0:
        sg.popup('No Assay Data for this product.')
        return
    elif statistics.mean(map(float, plotdfa1.A1)) != 0:
        try:
            specL1 = statistics.mean(map(float, plotdfa1.ASp1L))
            specH1 = statistics.mean(map(float, plotdfa1.ASp1H))

            w = 20
            h = 5
            d = 70
            plt.figure(figsize=(w, h), dpi=d)
            plt.xticks(rotation=75)
            plt.tick_params(axis='x', labelsize=10)
            plt.tick_params(bottom=True, top=False, left=True, right=True)
            plt.ylabel('% Wt. ' + str(plotdfa1.at[0, 'AC1_x']) + ' in ' + str(x1))
            plt.xlabel('Lot (Date Range: ' + str(datex.strftime('%b %d,%y')) + ' - ' + str(datetime.date.today().strftime('%b %d,%y')) + ')')
            plt.title('Titration History for ' + str(plotdfa1.iloc[0]['AC1_x']) + ' in ' + prodname + ' (' + str(PN) + ')')
    
            red_patch = mpatches.Patch(color='red', label='Control Limit +/- 2SD')
            blue_patch = mpatches.Patch(color='blue', label='Product Data')
            yellow_patchL = mpatches.Patch(color='orange', label='Product Spec Low: ' + str(specL1))
            yellow_patchH = mpatches.Patch(color='orange', label='Product Spec High: ' + str(specH1))
            purple_patch = mpatches.Patch(color='purple', label='Average % Wt.: ' + str(round(statistics.mean(plotdfa1.A1), 3)))
            white3 = mpatches.Patch(color='white', label='Filters: ' + str(specialname) + ' sorted by ' + str(sort))
            plt.legend(handles=[white3, red_patch, blue_patch, yellow_patchL, yellow_patchH, purple_patch]) 
            
    
            x = plotdfa1.Lot
    
            y1 = plotdfa1.A1
            
            plt.axhline(y=specL1, color='orange', linestyle='-')
            plt.axhline(y=specH1, color='orange', linestyle='-')     
            plt.axhline(y=statistics.mean(plotdfa1.A1), color='purple', linestyle='--')
            plt.axhline(y=statistics.mean(plotdfa1.A1)+(2*(statistics.stdev(plotdfa1.A1))), color='r', linestyle='-')
            plt.axhline(y=statistics.mean(plotdfa1.A1)-(2*(statistics.stdev(plotdfa1.A1))), color='r', linestyle='-')
            plt.plot(x, y1, marker='o')
            pathA1 = os.path.join(pathass, str(prodname) + ' ' + str(plotdfa1.at[0, 'AC1_x']) + ' ' + str(datex.strftime('%b %d,%y')) + '_' + str(datetime.date.today().strftime('%b %d,%y')) + '.png')
            plt.savefig(pathA1, bbox_inches='tight')
        
        except ValueError:
            sg.popup('Check product information and try again.')
            return
        
#plot A2
    if statistics.mean(map(float, plotdfa1.A2)) != 0:
        try:
            specL2 = statistics.mean(map(float, plotdfa1.ASp2L))
            specH2 = statistics.mean(map(float, plotdfa1.ASp2H))
        
            w = 20
            h = 5
            d = 70
            plt.figure(figsize=(w, h), dpi=d)
            plt.xticks(rotation=75)
            plt.tick_params(axis='x', labelsize=10)
            plt.tick_params(bottom=True, top=False, left=True, right=True)
            plt.ylabel('% Wt. ' + str(plotdfa1.at[0, 'AC2_x']) + ' in ' + str(x1))
            plt.xlabel('Lot (Date Range: ' + str(datex.strftime('%b %d,%y')) + ' - ' + str(datetime.date.today().strftime('%b %d,%y')) + ')')
            plt.title('Titration History for ' + str(plotdfa1.iloc[0]['AC2_x']) + ' in ' + prodname + ' (' + str(PN) + ')')
    
            red_patch = mpatches.Patch(color='red', label='Control Limit +/- 2SD')
            blue_patch = mpatches.Patch(color='blue', label='Product Data')
            yellow_patchL = mpatches.Patch(color='orange', label='Product Spec Low: ' + str(specL2))
            yellow_patchH = mpatches.Patch(color='orange', label='Product Spec High: ' + str(specH2))         
            purple_patch = mpatches.Patch(color='purple', label='Average % Wt.: ' + str(round(statistics.mean(plotdfa1.A2), 3)))
            white3 = mpatches.Patch(color='white', label='Filters: ' + str(specialname) + ' sorted by ' + str(sort))
            plt.legend(handles=[white3, red_patch, blue_patch, yellow_patchL, yellow_patchH, purple_patch]) 
    
            x = plotdfa1.Lot
    
            y1 = plotdfa1.A2
    
            plt.axhline(y=specL2, color='orange', linestyle='-')
            plt.axhline(y=specH2, color='orange', linestyle='-')
            plt.axhline(y=statistics.mean(plotdfa1.A2), color='purple', linestyle='--')
            plt.axhline(y=statistics.mean(plotdfa1.A2)+(2*(statistics.stdev(plotdfa1.A2))), color='r', linestyle='-')
            plt.axhline(y=statistics.mean(plotdfa1.A2)-(2*(statistics.stdev(plotdfa1.A2))), color='r', linestyle='-')
            plt.plot(x, y1, marker='o')
            pathA2 = os.path.join(pathass, str(prodname) + ' ' + str(plotdfa1.at[0, 'AC2_x']) + ' ' + str(datex.strftime('%b %d,%y')) + '_' + str(datetime.date.today().strftime('%b %d,%y')) + '.png')
            plt.savefig(pathA2, bbox_inches='tight')

        except ValueError:
            sg.popup('No 2nd or 3rd Assay Component')
            return
        
    elif statistics.mean(map(float, plotdfa1.A2)) == 0:
        pass
        
#plot A3
    if statistics.mean(map(float, plotdfa1.A3)) != 0:
        try:
            specL3 = statistics.mean(map(float, plotdfa1.ASp3L))
            specH3 = statistics.mean(map(float, plotdfa1.ASp3H))
    
            w = 20
            h = 5
            d = 70
            plt.figure(figsize=(w, h), dpi=d)
            plt.xticks(rotation=75)
            plt.tick_params(axis='x', labelsize=10)
            plt.tick_params(bottom=True, top=False, left=True, right=True)
            plt.ylabel('% Wt. ' + str(plotdfa1.at[0, 'AC3_x']) + ' in ' + str(x1))
            plt.xlabel('Lot (Date Range: ' + str(datex.strftime('%b %d,%y')) + ' - ' + str(datetime.date.today().strftime('%b %d,%y')) + ')')
            plt.title('Titration History for ' + str(plotdfa1.iloc[0]['AC3_x']) + ' in ' + prodname + ' (' + str(PN) + ')')
    
            red_patch = mpatches.Patch(color='red', label='Control Limit +/- 2SD')
            blue_patch = mpatches.Patch(color='blue', label='Product Data')
            yellow_patchL = mpatches.Patch(color='orange', label='Product Spec Low: ' + str(specL3))
            yellow_patchH = mpatches.Patch(color='orange', label='Product Spec High: ' + str(specH3))         
            purple_patch = mpatches.Patch(color='purple', label='Average % Wt.: ' + str(round(statistics.mean(plotdfa1.A3), 3)))
            white3 = mpatches.Patch(color='white', label='Filters: ' + str(specialname) + ' sorted by ' + str(sort))
            plt.legend(handles=[white3, red_patch, blue_patch, yellow_patchL, yellow_patchH, purple_patch])
            plt.axhline(y=statistics.mean(plotdfa1.A3), color='purple', linestyle='--')
    
            x = plotdfa1.Lot
    
            y1 = plotdfa1.A3
    
            plt.axhline(y=specL3, color='orange', linestyle='-')
            plt.axhline(y=specH3, color='orange', linestyle='-')
            plt.axhline(y=statistics.mean(plotdfa1.A3), color='purple', linestyle='--')
            plt.axhline(y=statistics.mean(plotdfa1.A3)+(2*(statistics.stdev(plotdfa1.A3))), color='r', linestyle='-')
            plt.axhline(y=statistics.mean(plotdfa1.A3)-(2*(statistics.stdev(plotdfa1.A3))), color='r', linestyle='-')
            plt.plot(x, y1, marker='o')
            pathA3 = os.path.join(pathass, str(prodname) + ' ' + str(plotdfa1.at[0, 'AC3_x']) + ' ' + str(datex.strftime('%b %d,%y')) + '_' + str(datetime.date.today().strftime('%b %d,%y')) + '.png')
            plt.savefig(pathA3, bbox_inches='tight')
        
        except ValueError:
            sg.popup('No 3rd Assay Component')
            
    elif statistics.mean(map(float, plotdfa1.A3)) == 0:
        pass

def pull1001assay(x1, sort, report, datex):
    data = pd.ExcelFile(workingpath)
    
    dfa = pd.DataFrame(pd.read_excel(data, 'InLabLogdata'))
    
    dfbc = pd.DataFrame(pd.read_excel(data,  'Spec'))
    dfb = dfbc[dfbc['Wking']=='X']
    
    dfaa = pd.merge(dfa, dfb, on = 'PN', how = "left")
    dfaa['ADT']= pd.to_datetime(dfaa['ADT'])
    dfaa['ADT']= dfaa['ADT'].dt.date
    
    plotdf= dfaa[(dfaa['ADT'] >= datex)]      
    
####
    if sort=='Lot':
        plotdf.sort_values('Lot', inplace = True)
        #plotdf.to_excel(r'C:\Users\HPP Assay\Desktop\Dump\plotdf.xlsx', sheet_name='1')     
    elif sort=='Date':
        plotdf.sort_values('ADT', inplace = True)
           
####              
    if type(x1)==str:
        plotdfa = plotdf[plotdf['Product_x']==x1]
    elif type(x1)==int:
        plotdfa = plotdf[plotdf['PN']==x1]
        #plotdfa.to_excel(r'C:\Users\HPP Assay\Desktop\Excel DUmp\plotdfa.xlsx', sheet_name='1')
    
####        
    if report =='Reported':
        is_reported = plotdfa['RA']=='X'
        plotdfa1 = plotdfa[is_reported]
        plotdfa1.reset_index(inplace = True)
        specialname = 'Reported Data'
        #plotdfa1.to_excel(r'C:\Users\HPP Assay\Desktop\Dump\plotdfa1.xlsx', sheet_name='1')
        #return
    
    elif report=='SPC':
        sg.popup('No SPC Data for this product.')
        return
        
    prodname = str(plotdfa1.iloc[0]['Product_x'])
    pna = round(plotdfa1.iloc[0]['PN'])
    PN = str(pna)
  
    specL1 = statistics.mean(map(float, plotdfa1.ASp1L))
    specH1 = statistics.mean(map(float, plotdfa1.ASp1H))
    
    avet=round(statistics.mean(plotdfa1.A1), 3)
    aveb=round(statistics.mean(plotdfa1.A3), 3)
    
    CL1H = statistics.mean(plotdfa1.A1)+(2*(statistics.stdev(plotdfa1.A1)))
    CL1L = statistics.mean(plotdfa1.A1)-(2*(statistics.stdev(plotdfa1.A1)))
    CL2H = statistics.mean(plotdfa1.A3)+(2*(statistics.stdev(plotdfa1.A3)))
    CL2L = statistics.mean(plotdfa1.A3)-(2*(statistics.stdev(plotdfa1.A3)))
    CLH = (CL1H+CL2H)/2
    CLL = (CL1L+CL2L)/2
    
    w = 20
    h = 7
    d = 70
    plt.figure(figsize=(w, h), dpi=d)
    plt.xticks(rotation=75)
    plt.tick_params(axis='x', labelsize=10)
    plt.tick_params(bottom=True, top=False, left=True, right=True)
    plt.ylabel('% Wt. ' + str(plotdfa1.at[0, 'AC1_x']) + ' in ' + str(x1))
    plt.xlabel('Lot (Date Range: ' + str(datex.strftime('%b %d,%y')) + ' - ' + str(datetime.date.today().strftime('%b %d,%y')) + ')')
    plt.title('Titration History for ' + str(plotdfa1.iloc[0]['AC1_x']) + ' in ' + prodname + ' (' + str(PN) + ')')

    red_patch = mpatches.Patch(color='red', label='Control Limit +/- 2SD (All)')
    blue_patch = mpatches.Patch(color='blue', label='Tote Data')
    green_patch = mpatches.Patch(color='green', label='BOR Data')
    yellow_patchL = mpatches.Patch(color='orange', label='Product Spec Low: ' + str(specL1))
    yellow_patchH = mpatches.Patch(color='orange', label='Product Spec High: ' + str(specH1))
    white3 = mpatches.Patch(color='white', label='Filters: ' + str(specialname) + ' sorted by ' + str(sort))
    white1 = mpatches.Patch(color='white', label='Average % Wt. (Tote): ' + str(avet))
    white2 = mpatches.Patch(color='white', label='Average % Wt. (BOR): ' + str(aveb))
    plt.legend(handles=[white3, white1, white2, red_patch, blue_patch, green_patch, yellow_patchL, yellow_patchH]) 

    x = plotdfa1.Lot

    y1 = plotdfa1.A1
    y2 = plotdfa1.A3
    
    plt.axhline(y=specL1, color='orange', linestyle='-')
    plt.axhline(y=specH1, color='orange', linestyle='-')     
    plt.axhline(y=statistics.mean(plotdfa1.A1), color='blue', linestyle='--')
    plt.axhline(y=statistics.mean(plotdfa1.A3), color='green', linestyle='--')
    plt.axhline(y=CLH, color='r', linestyle='-')
    plt.axhline(y=CLL, color='r', linestyle='-')
    plt.plot(x, y1, marker='o', color = 'blue')
    plt.plot(x, y2, marker='o', color = 'green')
    pathA1 = os.path.join(pathass, str(prodname) + ' ' + str(plotdfa1.at[0, 'AC1_x']) + ' ' + str(datex.strftime('%b %d,%y')) + '_' + str(datetime.date.today().strftime('%b %d,%y')) + '.png')
    plt.savefig(pathA1, bbox_inches='tight')

    
def assayview():
    layout= [
            [sg.Text('Assay Data Viewer')],
            [sg.Text('Choose a your plot below:')],
            [sg.Listbox(values=fnames, enable_events=True, size=(60,20),key='-FILE LIST-')],
            [sg.Image(key='-IMAGE-')],
            ]
    return sg.Window('HPP Lab Data Manager', layout)

def partsview():
    layout= [
            [sg.Text('Particle Data Viewer')],
            [sg.Text('Choose a your plot below:')],
            [sg.Listbox(values=fnames, enable_events=True, size=(60,20),key='-FILE LIST-')],
            [sg.Image(key='-IMAGE-')],
            ]
    return sg.Window('HPP Lab Data Manager', layout)

def DIWview():
    layout= [
            [sg.Text('DIW Data Viewer')],
            [sg.Text('Choose a your plot below:')],
            [sg.Listbox(values=fnames, enable_events=True, size=(60,20),key='-FILE LIST-')],
            [sg.Image(key='-IMAGE-')],
            ]
    return sg.Window('HPP Lab Data Manager', layout)

def preroot():
    layout= [         
            [sg.Text('Welcome to the HPP Data Manager!')],
            [sg.Text("  ")],
            [sg.Text('Open the "Lab Data" file using the "Browse" button')],
            [sg.Input(), sg.FileBrowse()],
            [sg.Button('Continue')]
            ]
    return sg.Window('HPP Lab Data Manager', layout)
    
def rootlayout():
    layout= [         
            [sg.Text('HPP Lab Data Manager')],
            [sg.Text("  ")],
            [sg.Text('Lab Data File:'), sg.Text(str(workingpath[-13:]), text_color = 'red'), sg.Button('Change')],
            [sg.Text("  ")],
            [sg.Text('Report Tools:')],
            [sg.Button('Reports'), sg.Button('Counter'), sg.Button('DIW')],
            [sg.Text('Data Tracking:')],
            [sg.Button('Particles'), sg.Button('Assays')]
            ]
    return sg.Window('HPP Lab Data Manager', layout)

def reportslayout():
    layout= [
            [sg.Text('Product Reporting Tool')],
            [sg.Text(' ')],
            [sg.Text('Choose where to generate Reports using the "Browse" button.')],
            [sg.Input(), sg.FolderBrowse()],
            [sg.OK()]
            ]
    return sg.Window('HPP Lab Data Manager', layout)

def DIW1layout():
    layout= [
            [sg.Text('DIW Tracker')],
            [sg.Text(' ')],
            [sg.Text('Generate DIW Quality Plot for:'), sg.Button('Previous 3 Months'), ],
            [sg.Text(' ')],
            [sg.Text('To generate a custom DIW Quality Plot, choose a:')],
            [sg.Button('Start Date')]
            ]
    return sg.Window('HPP Lab Data Manager', layout)

def DIW2layout():
    layout= [
            [sg.Text('DIW Tracker')],
            [sg.Text('Start Date:'), sg.Text(str(Date1a.strftime('%b %d, %Y')), text_color = 'red'), sg.Button('Change')],
            [sg.Text(' ')],
            [sg.Text('Choose End Date:')],
            [sg.Button('End Date')]
            ]
    return sg.Window('HPP Lab Data Manager', layout)

def DIW3layout():
    layout= [
            [sg.Text('Do you want to delete existing DIW Plots?')],
            [sg.Button('Yes'), sg.Button('No')]
            ]
    return sg.Window('HPP Lab Data Manager', layout)

def counter1layout():
    layout= [
            [sg.Text('HPP Lab Counter')],
            [sg.Text(' ')],
            [sg.Text('Generate Counter Report for:'), sg.Button('Previous Month')],
            [sg.Text(' ')],
            [sg.Text('To generate a custom Counter Report, choose a:')],
            [sg.Button('Start Date')]
            ]
    return sg.Window('HPP Lab Data Manager', layout)

def counter2layout():
    layout= [
            [sg.Text('HPP Lab Counter')],
            [sg.Text(' ')],
            [sg.Text('Start Date:'), sg.Text(str(Date1a.strftime('%b %d, %Y')), text_color = 'red'), sg.Button('Change')],
            [sg.Text(' ')],
            [sg.Text('Choose End Date:')],
            [sg.Button('End Date')]
            ]
    return sg.Window('HPP Lab Data Manager', layout)

def counter3layout():
    layout= [
            [sg.Text('Would you like to generate an excel file "Counter Report"?')],
            [sg.Button('Yes'), sg.Button('No')]
            ]
    return sg.Window('HPP Lab Data Manager', layout)

def particles1layout():
    layout= [
            [sg.Text('Particle Data Tracking',)],
            [sg.Text(' ')],
            [sg.Text('Enter Product Number:'), sg.Input(b)],
            [sg.Text('Choose a Start Date:'), sg.Text(str(a), text_color = 'red'), sg.Button('Start Date')],
            [sg.Text('Data Selection:')],
            [sg.Radio('Reported Data', 'Beta'), sg.Radio('All Data', 'Beta')],
            [sg.Text('Sort Options:')],
            [sg.Radio('Date', 'Alpha'), sg.Radio('Lot #', 'Alpha')],
            [sg.Button('Plot')],
            ]
    return sg.Window('HPP Lab Data Manager', layout)

def particles2layout():
    layout= [
            [sg.Text('Do you want to delete existing Particle Plots?')],
            [sg.Button('Yes'), sg.Button('No')]
            ]
    return sg.Window('HPP Lab Data Manager', layout)

def assay1layout():
    layout= [
            [sg.Text('Assay Data Tracking')],
            [sg.Text(' ')],
            [sg.Text('Enter Product Name or Number:'), sg.Input(b)],
            [sg.Text('Choose a Start Date:'), sg.Text(str(a), text_color = 'red'), sg.Button('Start Date')],
            [sg.Text('Data Selection:')],
            [sg.Radio('Reported Data', 'Beta'), sg.Radio('SPC Data', 'Beta')],
            [sg.Text('Sort Options:')],
            [sg.Radio('Date', 'Alpha'), sg.Radio('Lot #', 'Alpha')],
            [sg.Button('Plot')]
            ]
    return sg.Window('HPP Lab Data Manager', layout)

def assay2layout():
    layout= [
            [sg.Text('Do you want to delete existing Assay Plots?')],
            [sg.Button('Yes'), sg.Button('No')]
            ]
    return sg.Window('HPP Lab Data Manager', layout)

#GUI STRUCTURE
          
while True: #preroot window
    windowpre = preroot() #This window gets the lab data file
    events, valuess = windowpre.read()
    
    if events == sg.WIN_CLOSED or events == windowpre.close():
        windowpre.close()
        break
    
    workingpath = valuess[0] #Lab Data file
    
    if events == 'Continue':
        if workingpath == '':
            sg.popup('You must choose the "Lab Data" file to continue.')
        else:
            break
        
while True: #Main Menu Starts here
    window = rootlayout() #Home window everthing out of here
    event, values = window.read()
    
    if event == sg.WIN_CLOSED or event == window.close():
            window.close()
            break
        
    if event == 'Change':
       while True:
            windowpre = preroot() #This window gets the lab data file
            events, valuess = windowpre.read()
            
            if events == sg.WIN_CLOSED or events == windowpre.close():
                windowpre.close()
                break
            
            workingpath = valuess[0] #Lab Data file
            
            if events == 'Continue':
                if workingpath == '':
                    sg.popup('You must choose the "Lab Data" file to continue.')
                else:
                    break
    
    if event == 'Reports':
        while True:
            window2 = reportslayout() #Generates daily lab analytical work reports.
            a, valuesa = window2.read()
            
            if a == sg.WIN_CLOSED or a == window2.close() or event == window.close():
                window2.close()
                break
        
            Path2 = valuesa[0] #file where reports are generated
            
            if Path2 != '':
                report(Path2)
                window2.close()
                break
            else:
                sg.popup('Choose File for report generation.')
                continue
            
    if event == 'Counter':                   
        while True:
            window3 = counter1layout() #counter for monthly reports
            b, valuesb = window3.read()
            
            if b == sg.WIN_CLOSED or b == window3.close() or event == window.close():
                window3.close()
                break
            
            if b == 'Previous Month':
                today = datetime.date.today()
                first = today.replace(day=1)
                lastdayprevMonth = first - datetime.timedelta(days=1)
                firstdayprevMonth = lastdayprevMonth.replace(day=1)
                Date1a = datetime.datetime.strptime(str(firstdayprevMonth), '%Y-%m-%d').date()
                Date2a = datetime.datetime.strptime(str(lastdayprevMonth), '%Y-%m-%d').date()
                
                while True: 
                    window3b = counter3layout() #option to generate an excel report 
                    bb, valuesbb = window3b.read()
                    
                    if bb == sg.WIN_CLOSED or bb == window3b.close() or event == window.close():
                        window3b.close()
                        break
                    
                    if bb == 'Yes':
                        window3b.close()
                        desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
                        path=os.path.join(desktop, 'Lab Plots')
                        try:                        #Attempts to make the file if it does not exist
                            os.makedirs(path)
                        except FileExistsError:
                            pass

                        counter(Date1a, Date2a)
                        sg.popup('Report are in the "Lab Plots" folder on your Desktop.', background_color = 'Light Blue')
                        break
                
                    if bb == 'No':
                        window3b.close()
                        counter(Date1a, Date2a)
                        break
                break
                
            if b == 'Start Date':
                try:
                    x = sg.popup_get_date() #chooses start date from calendar and then formats it appropriately
                    starty = int(x[2])
                    startm = int(x[0])
                    startd = int(x[1])
                    startdate = str(starty) + '-' + str(startm) + '-' + str(startd)
                    Date1a = datetime.datetime.strptime(startdate, '%Y-%m-%d').date()
                except TypeError:
                    continue
                
                if  Date1a > datetime.date.today():
                    sg.popup('Choose Start Date Before: ' + str(datetime.date.today()))
                    continue
            
                while True:
                    window3a = counter2layout() #opens window with option to change start date and choose and end date
                    ba, valuesba = window3a.read()
                    
                    if ba == sg.WIN_CLOSED or ba == window3a.close() or event == window.close():
                        window3a.close()
                        break
                    
                    if ba == 'Change':
                        break
                        
                    if ba == 'End Date':
                        window3a.close()
                        try:
                            y = sg.popup_get_date() #Date pick and format
                            endy = int(y[2])
                            endm = int(y[0])
                            endd = int(y[1])
                            enddate = str(endy) + '-' + str(endm) + '-' + str(endd)
                            Date2a = datetime.datetime.strptime(enddate, '%Y-%m-%d').date()
                        except TypeError:
                            continue
                        
                        if  Date2a > datetime.date.today():
                            sg.popup('Choose End Date Before: ' + str(datetime.date.today()))
                            continue
                        
                        try:
                            if Date1a >= Date2a:
                                sg.popup('End Date must be after Start Date')
                                continue
                            sg.popup('End Date: ' + str(Date2a.strftime('%b %d, %Y')))
                        except NameError:
                            sg.popup('Choose a Start Date First')
                            continue
                        
                        while True: 
                            window3b = counter3layout() #option to generate an excel report 
                            bb, valuesbb = window3b.read()
                            
                            if bb == sg.WIN_CLOSED or bb == window3b.close() or event == window.close():
                                window3b.close()
                                break
                            
                            if bb == 'Yes':
                                window3b.close()
                                desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
                                path=os.path.join(desktop, 'Lab Plots')
                                try:                        #Attempts to make the file if it does not exist
                                    os.makedirs(path)
                                except FileExistsError:
                                    pass

                                counter(Date1a, Date2a)
                                sg.popup('Report can befound in the "Lab Plots" folder on your Desktop.', background_color = 'Light Blue')
                                break
                                
                            if bb == 'No':
                                window3b.close()
                                counter(Date1a, Date2a)
                                break
                        break
    
    if event == 'DIW':
        while True:
            window6 = DIW1layout()
            e, valuese = window6.read()
            
            if e == sg.WIN_CLOSED or e == window6.close() or event == window.close():
                window6.close()
                break
            
            if e == 'Previous 3 Months':
                today = datetime.date.today()
                first = today.replace(day=1)
                lastdayprevMonth = first - datetime.timedelta(days=1)
                firstdayprevMonth = lastdayprevMonth.replace(day=1)
                minusthreefive = firstdayprevMonth - datetime.timedelta(days=35)
                firstday3Month = minusthreefive.replace(day=1)
                Date1a = datetime.datetime.strptime(str(firstday3Month), '%Y-%m-%d').date()
                Date2a = datetime.datetime.strptime(str(lastdayprevMonth), '%Y-%m-%d').date()
                
                desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
                path=os.path.join(desktop, 'Lab Plots')
                pathDIW=os.path.join(path, 'DIW Plots')
                
                while True:
                    window4a = DIW3layout()
                    ca, valuesca = window4a.read()
            
                    if ca == sg.WIN_CLOSED or ca == window6.close() or event == window.close():
                        window4a.close()
                        break
                    
                    if ca == 'Yes':
                        try:
                            shutil.rmtree(pathDIW)
                            sg.popup('DIW Plots deleted.')
                            window4a.close()
                            break
                        except:
                            sg.popup('DIW Plots deleted.')
                            window4a.close()
                            break
                    if ca == 'No':
                        window4a.close()
                        break
                    
                try:
                    os.makedirs(path)
                except FileExistsError:
                    pass

                try:
                    os.makedirs(pathDIW)
                except FileExistsError:
                    pass
                
                try:
                    DIW(Date1a, Date2a)
                    folder = pathDIW
                    file_list = os.listdir(folder)         # get list of files in folder
                    fnames = [f for f in file_list if os.path.isfile(
                        os.path.join(folder, f)) and f.lower().endswith((".png", ".jpg", "jpeg", ".tiff", ".bmp"))]
                    
                    window5b = DIWview()
                    while True:
                        event, values = window5b.read()
                        if event in (sg.WIN_CLOSED, 'Exit'):
                            break
                        
                        if event == sg.WIN_CLOSED or event == 'Exit':
                            break
                        
                        if event == '-FILE LIST-':    # A file was chosen from the listbox
                            try:
                                filename = os.path.join(folder, values['-FILE LIST-'][0])
                                window5b['-IMAGE-'].update(data=convert_to_bytes(filename))
                                continue
                            except IndexError:
                                window5b.close()
                                break
                            
                    sg.popup('All plots can be found in the "Lab Plots" folder on your Desktop', background_color = 'Light Blue')
                    break
                except:
                    sg.popup('Fix formating in Lab Data sheet.')
                    break
                
            if e == 'Start Date':
                try:
                    x = sg.popup_get_date()
                    starty = int(x[2])
                    startm = int(x[0])
                    startd = int(x[1])
                    startdate = str(starty) + '-' + str(startm) + '-' + str(startd)
                    Date1a = datetime.datetime.strptime(startdate, '%Y-%m-%d').date()
                except TypeError:
                    continue
                
                if  Date1a > datetime.date.today():
                    sg.popup('Choose Start Date Before: ' + str(datetime.date.today()))
                    continue
                
                while True:
                    window6a = DIW2layout()
                    ea, valuesea = window6a.read()
                    
                    if ea == sg.WIN_CLOSED or ea == window6a.close() or event == window.close():
                        window6a.close()
                        break
                    
                    if ea == 'Change':
                        break
                    
                    if ea == 'End Date':
                        try:
                            y = sg.popup_get_date()
                            endy = int(y[2])
                            endm = int(y[0])
                            endd = int(y[1])
                            enddate = str(endy) + '-' + str(endm) + '-' + str(endd)
                            Date2a = datetime.datetime.strptime(enddate, '%Y-%m-%d').date()
                        except TypeError:
                            continue
                        
                        if  Date2a > datetime.date.today():
                            sg.popup('Choose End Date Before: ' + str(datetime.date.today()))
                            continue
                        
                        try:
                            if Date1a >= Date2a:
                                sg.popup('End Date must be after Start Date')
                                continue
                            sg.popup('End Date: ' + str(Date2a.strftime('%b %d, %Y')))
                        except NameError:
                            sg.popup('Choose a Start Date First')
                            continue
                        
                        desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
                        path=os.path.join(desktop, 'Lab Plots')
                        pathDIW=os.path.join(path, 'DIW Plots')
                        
                        while True:
                            window4a = DIW3layout()
                            ca, valuesca = window4a.read()
                    
                            if ca == sg.WIN_CLOSED or ca == window6.close() or event == window.close():
                                window4a.close()
                                break
                            
                            if ca == 'Yes':
                                try:
                                    shutil.rmtree(pathDIW)
                                    sg.popup('DIW Plots deleted.')
                                    window4a.close()
                                    break
                                except:
                                    sg.popup('DIW Plots deleted.')
                                    window4a.close()
                                    break
                            if ca == 'No':
                                window4a.close()
                                break
                            
                        try:
                            os.makedirs(path)
                        except FileExistsError:
                            pass

                        try:
                            os.makedirs(pathDIW)
                        except FileExistsError:
                            pass
                        
                        try:
                            DIW(Date1a, Date2a)
                            folder = pathDIW
                            file_list = os.listdir(folder)         # get list of files in folder
                            fnames = [f for f in file_list if os.path.isfile(
                                os.path.join(folder, f)) and f.lower().endswith((".png", ".jpg", "jpeg", ".tiff", ".bmp"))]
                            
                            window5b = DIWview()
                            while True:
                                event, values = window5b.read()
                                if event in (sg.WIN_CLOSED, 'Exit'):
                                    break
                                
                                if event == sg.WIN_CLOSED or event == 'Exit':
                                    break
                                
                                if event == '-FILE LIST-':    # A file was chosen from the listbox
                                    try:
                                        filename = os.path.join(folder, values['-FILE LIST-'][0])
                                        window5b['-IMAGE-'].update(data=convert_to_bytes(filename))
                                        continue
                                    except IndexError:
                                        window5b.close()
                                        break
                                    
                            sg.popup('All plots can be found in the "Lab Plots" folder on your Desktop', background_color = 'Light Blue')
                            break
                        except ValueError or TypeError:
                            sg.popup('Fix date formating in Lab Data sheet.')
                            break
                                    
    if event == 'Particles':
        a = ''
        b = ''
        while True:
            window4 = particles1layout()
            c, valuesc = window4.read()
            
            if c == sg.WIN_CLOSED or c == window4.close() or event == window.close():
                window4.close()
                break
            
            ProdNum = valuesc[0]
            b = ProdNum
            desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            path=os.path.join(desktop, 'Lab Plots')
            pathpart=os.path.join(path, 'Particle Plots')
            
                
            if c == 'Start Date':
                try:    
                    aa = sg.popup_get_date()
                    starty = int(aa[2])
                    startm = int(aa[0])
                    startd = int(aa[1])
                    startdate = str(starty) + '-' + str(startm) + '-' + str(startd)
                    a = datetime.datetime.strptime(startdate, '%Y-%m-%d').date()
                except:
                    break
                
                if  a >= datetime.date.today():
                    sg.popup('Choose Start Date Before: ' + str(datetime.date.today()))
                    
                
            if c == 'Plot':
                if ProdNum == '':
                    sg.popup('Enter a Product Number')
                    continue
                if a == '':
                    sg.popup('Choose a Start Date')
                    continue
                if valuesc[1] == False and valuesc[2] == False:
                    sg.popup('Select data filters before proceeding.')
                    continue
                if valuesc[3] == False and valuesc[4] == False:
                    sg.popup('Select data filters before proceeding.')
                    continue
                
                try:
                    x=int(ProdNum)
                except ValueError:
                    sg.popup("Invalid Product Number!")
                    continue
                if valuesc[1]==True:
                    z='Reported'
                elif valuesc[1]==False:
                    z='All'
                
                if valuesc[3]==True:    
                    y='Date'
                elif valuesc[3]==False:
                    y='Lot'
                
                if valuesc[1]==False and valuesc[3]== True:
                    sg.popup("Must sort by 'Lot' when using the 'All' argument.")
                    continue
                
                while True:
                    window4a = particles2layout()
                    ca, valuesca = window4a.read()
            
                    if ca == sg.WIN_CLOSED or ca == window4.close() or event == window.close():
                        window4a.close()
                        break
                    
                    if ca == 'Yes':
                        try:
                            shutil.rmtree(pathpart)
                            sg.popup('Particle Plots deleted.')
                            window4a.close()
                            break
                        except:
                            sg.popup('Particle Plots deleted.')
                            window4a.close()
                            break
                    if ca == 'No':
                        window4a.close()
                        break
    
                try:
                    os.makedirs(path)
                except FileExistsError:
                    pass
    
                try:
                    os.makedirs(pathpart)
                except FileExistsError:
                    pass
                    
                try:
                    pullparts(x,y,z,a)
                except:
                    sg.popup('Check Product Number\nOr... \nExpand Date Range')
                    continue
                
                folder = pathpart
                file_list = os.listdir(folder)         # get list of files in folder
                fnames = [f for f in file_list if os.path.isfile(
                    os.path.join(folder, f)) and f.lower().endswith((".png", ".jpg", "jpeg", ".tiff", ".bmp"))]
                
                window5b = partsview()
                while True:
                    event, values = window5b.read()
                    if event in (sg.WIN_CLOSED, 'Exit'):
                        break
                    
                    if event == sg.WIN_CLOSED or event == 'Exit':
                        break
                    
                    if event == '-FILE LIST-':    # A file was chosen from the listbox
                        try:
                            filename = os.path.join(folder, values['-FILE LIST-'][0])
                            window5b['-IMAGE-'].update(data=convert_to_bytes(filename))
                            continue
                        except IndexError:
                            window5b.close()
                            break
                break
                                                                    
    if event == 'Assays':
        a = ''
        b = ''
        while True:
            window5 = assay1layout()
            d, valuesd = window5.read()
            
            if d == sg.WIN_CLOSED or d == window5.close() or event == window.close():
                window5.close()
                break
            
            PrdN = valuesd[0]
            b = PrdN
            desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            path=os.path.join(desktop, 'Lab Plots')
            pathass=os.path.join(path, 'Assay Plots')
            
            if d == 'Start Date':
                try:    
                    aa = sg.popup_get_date()
                    starty = int(aa[2])
                    startm = int(aa[0])
                    startd = int(aa[1])
                    startdate = str(starty) + '-' + str(startm) + '-' + str(startd)
                    a = datetime.datetime.strptime(startdate, '%Y-%m-%d').date()
                except:
                    break
                
                if  a >= datetime.date.today():
                    sg.popup('Choose Start Date Before: ' + str(datetime.date.today()))
            
            if d=='Plot':
                if PrdN == '':
                    sg.popup('Enter a Product Number')
                    continue
                if a == '':
                    sg.popup('Choose a Start Date')
                    continue
                if valuesd[1] == False and valuesd[2] == False:
                    sg.popup('Select data filters before proceeding.')
                    continue
                if valuesd[3] == False and valuesd[4] == False:
                    sg.popup('Select data filters before proceeding.')
                    continue
                
                try:
                    x=int(PrdN)
                except ValueError:
                    x=str(PrdN)

                if valuesd[1]==True:
                    z='Reported'
                elif valuesd[1]==False:
                    z='SPC'
                
                if valuesd[3]==True:    
                    y='Date'
                elif valuesd[3]==False:
                    y='Lot'
                
                if valuesd[1]==False and valuesd[3]== True:
                    sg.popup("Must sort by 'Lot' when using the 'SPC' argument.")  
                    continue
                
                desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
                path=os.path.join(desktop, 'Lab Plots')
                pathass=os.path.join(path, 'Assay Plots')
                
                while True:
                    window4a = assay2layout()
                    ca, valuesca = window4a.read()
            
                    if ca == sg.WIN_CLOSED or ca == window5.close() or event == window.close():
                        window4a.close()
                        break
                    
                    if ca == 'Yes':
                        try:
                            shutil.rmtree(pathass)
                            sg.popup('Assay Plots deleted.')
                            window4a.close()
                            break
                        except:
                            sg.popup('Assay Plots deleted.')
                            window4a.close()
                            break
                    if ca == 'No':
                        window4a.close()
                        break
                
                try:
                    os.makedirs(path)
                except FileExistsError:
                    pass

                try:
                    os.makedirs(pathass)
                except FileExistsError:
                    pass
                
                if x == 301239 or x == 300959 or x == 'Hydrofluoric Acid 100-1' or x == 'Hydrofluoric Acid 0.5% (100-1)' or x == 300049:
                    try:
                        pull1001assay(x,y,z,a)
                    except:
                        sg.popup('Check Product Name/Number\nOr... \nExpand Date Range')
                        continue
                else:
                    try:
                        pullassay(x,y,z,a)
                    except ValueError:
                        sg.popup('Check Product Name/Number\nOr... \nExpand Date Range')
                        continue
                    except IndexError:
                        sg.popup('Check Filter options. Most products do not have SPC data.')
                        continue
                
                folder = pathass
                file_list = os.listdir(folder)         # get list of files in folder
                fnames = [f for f in file_list if os.path.isfile(
                    os.path.join(folder, f)) and f.lower().endswith((".png", ".jpg", "jpeg", ".tiff", ".bmp"))]
                
                window5b = assayview()
                while True:
                    event, values = window5b.read()
                    if event in (sg.WIN_CLOSED, 'Exit'):
                        break
                    
                    if event == sg.WIN_CLOSED or event == 'Exit':
                        break
                    
                    if event == '-FILE LIST-':    # A file was chosen from the listbox
                        try:
                            filename = os.path.join(folder, values['-FILE LIST-'][0])
                            window5b['-IMAGE-'].update(data=convert_to_bytes(filename))
                            continue
                        except IndexError:
                            window5b.close()
                            break
                break


                
